from fastapi import FastAPI, Depends, HTTPException, File, UploadFile, Query
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import create_engine, Column, Integer, String, Float, Date, DateTime, ForeignKey, func, BigInteger, Boolean
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, Session
from passlib.context import CryptContext
from jose import JWTError, jwt
from datetime import datetime, timedelta, date as date_type
from typing import Optional, List
from dotenv import load_dotenv
from secrets import token_urlsafe
from urllib.parse import quote
from sqlalchemy.pool import NullPool
from sqlalchemy import DateTime
from datetime import date


import hashlib
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import io
import csv


# Load environment variables FIRST before importing custom modules
load_dotenv()

from fastapi import APIRouter
import requests
from urllib.parse import urlencode

# Import SMS parser router (after load_dotenv!)
from sms_parser_api import router as sms_router

# Import voice transaction router (after load_dotenv!)
from voice_transaction_api import router as voice_router

import json

from fastapi.responses import RedirectResponse

# Configuration
SECRET_KEY = os.getenv("SECRET_KEY", "your-secret-key-change-in-production-PLEASE")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.getenv("ACCESS_TOKEN_EXPIRE_MINUTES", "36500000"))
DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///./expense_tracker.db")
FRONTEND_URL = os.getenv("FRONTEND_URL", "http://localhost:5173")
API_BASE = os.getenv("API_BASE_URL", "https://webapp-expense.onrender.com")


SPLITWISE_CLIENT_ID = os.getenv("SPLITWISE_CLIENT_ID")
SPLITWISE_CLIENT_SECRET = os.getenv("SPLITWISE_CLIENT_SECRET")
SPLITWISE_REDIRECT_URI = os.getenv(
    "SPLITWISE_REDIRECT_URI",
    f"{API_BASE}/api/splitwise/callback"
)
SPLITWISE_BASE_URL = "https://secure.splitwise.com"


# 📧 Email Configuration
SMTP_HOST = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "")
FROM_EMAIL = os.getenv("FROM_EMAIL", os.getenv("SMTP_USER", "noreply@expensetracker.com"))
FROM_NAME = os.getenv("FROM_NAME", "Expense Tracker")
EMAIL_ENABLED = bool(SMTP_USER and SMTP_PASSWORD)


# Fix postgres:// to postgresql:// for SQLAlchemy
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

# Optimize for Supabase
if "supabase.com" in DATABASE_URL:
    # Use Session Pooler for better Render compatibility
    if "?" not in DATABASE_URL:
        DATABASE_URL += "?sslmode=require"
    
    print("🔧 Using Supabase with optimized settings")

# Create engine with proper configuration
if DATABASE_URL.startswith("sqlite"):
    engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
else:
    engine = create_engine(
    DATABASE_URL,
    pool_size=5,
    max_overflow=10,
    pool_pre_ping=True,
    pool_recycle=300,
)


print(f"📍 Database: {DATABASE_URL.split('@')[1].split('/')[0] if '@' in DATABASE_URL else 'sqlite'}")

SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()
Base.metadata.schema = "public"

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

# ==========================================
# DATABASE MODELS
# ==========================================

class User(Base):
    __tablename__ = "users"
    id = Column(Integer, primary_key=True, index=True)
    username = Column(String, unique=True, index=True)
    email = Column(String, unique=True, index=True)
    hashed_password = Column(String)
    
    # Profile fields
    full_name = Column(String, nullable=True)
    phone = Column(String, nullable=True)
    date_of_birth = Column(Date, nullable=True)
    occupation = Column(String, nullable=True)
    monthly_budget = Column(Float, nullable=True)
    onboarding_completed = Column(Boolean, default=False)  # SQLite uses 0/1 for boolean
    
    # Timestamps
    created_at = Column(DateTime, default=datetime.utcnow)

      # 🔹 Splitwise integration
    splitwise_user_id = Column(Integer, nullable=True)           # your SW user ID
    splitwise_access_token = Column(String, nullable=True)
    splitwise_refresh_token = Column(String, nullable=True)
    splitwise_token_expires_at = Column(DateTime, nullable=True)
    splitwise_last_sync_at = Column(DateTime, nullable=True)

class PasswordResetToken(Base):
    __tablename__ = "password_reset_tokens"

    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, ForeignKey("users.id"))
    token = Column(String, unique=True, index=True)

    created_at = Column(DateTime, default=datetime.utcnow)
    expires_at = Column(DateTime, nullable=False)
    used = Column(Integer, default=0)

class PendingTransaction(Base):
    __tablename__ = "pending_transactions"
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, ForeignKey("users.id"))
    token = Column(String, unique=True, index=True)
    amount = Column(Float, nullable=True)
    category = Column(String, nullable=True)
    description = Column(String, nullable=True)
    date = Column(String, nullable=True)
    type = Column(String, nullable=True)
    created_at = Column(Date, default=datetime.utcnow)
    status = Column(String, default="pending")
     # 🔹 Splitwise-specific
    splitwise_expense_id = Column(BigInteger, nullable=True, index=True)
    splitwise_group_name = Column(String, nullable=True)
    splitwise_raw_json = Column(String, nullable=True)  # optional, for debugging

class Expense(Base):
    __tablename__ = "expenses"
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, ForeignKey("users.id"))
    amount = Column(Float)
    category = Column(String)
    description = Column(String)
    date = Column(Date)
    type = Column(String)

# ✅ NEW: Category Model
class Category(Base):
    __tablename__ = "categories"
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, ForeignKey("users.id"), nullable=False)
    name = Column(String(50), nullable=False)
    color = Column(String(7), default="#667EEA")
    icon = Column(String(50), default="📦")
    created_at = Column(DateTime, default=datetime.utcnow)

# Create tables
# Base.metadata.create_all(bind=engine)

# ==========================================
# PYDANTIC MODELS
# ==========================================

class UserCreate(BaseModel):
    username: str
    email: str
    password: str

class UserLogin(BaseModel):
    username: str
    password: str

class ForgotPasswordRequest(BaseModel):
    email: str

class ResetPasswordRequest(BaseModel):
    token: str
    new_password: str

class Token(BaseModel):
    access_token: str
    token_type: str

class ExpenseCreate(BaseModel):
    amount: float
    category: str
    description: str
    date: str
    type: str

class ExpenseUpdate(BaseModel):
    amount: Optional[float] = None
    category: Optional[str] = None
    description: Optional[str] = None
    date: Optional[str] = None
    type: Optional[str] = None

class ExpenseResponse(BaseModel):
    id: int
    amount: float
    category: str
    description: Optional[str] = None
    date: date
    type: str
    
    class Config:
        from_attributes = True

class PendingTransactionResponse(BaseModel):
    id: int
    token: str
    amount: Optional[float]
    category: Optional[str]
    description: Optional[str]
    date: Optional[str]
    type: Optional[str]
    status: str
    
    class Config:
        from_attributes = True

class PendingTransactionCreate(BaseModel):
    amount: Optional[float] = None
    category: Optional[str] = None
    description: Optional[str] = None
    date: Optional[str] = None
    type: Optional[str] = None

# ✅ NEW: Category Pydantic Models
class CategoryCreate(BaseModel):
    name: str
    color: Optional[str] = "#667EEA"
    icon: Optional[str] = "📦"

class CategoryUpdate(BaseModel):
    name: Optional[str] = None
    color: Optional[str] = None
    icon: Optional[str] = None

class CategoryResponse(BaseModel):
    id: int
    name: str
    color: str
    icon: str
    created_at: datetime
    
    class Config:
        from_attributes = True


class UserProfile(BaseModel):
    full_name: Optional[str] = None
    phone: Optional[str] = None
    date_of_birth: Optional[str] = None
    occupation: Optional[str] = None
    monthly_budget: Optional[float] = None

class UserProfileResponse(BaseModel):
    username: str
    email: str
    full_name: Optional[str]
    phone: Optional[str]
    date_of_birth: Optional[str]
    occupation: Optional[str]
    monthly_budget: Optional[float]
    onboarding_completed: bool
    created_at: Optional[datetime]
    
    # Splitwise integration fields
    splitwise_user_id: Optional[int] = None
    splitwise_last_sync_at: Optional[datetime] = None
    
    class Config:
        from_attributes = True

class SignupWithProfile(BaseModel):
    username: str
    email: str
    password: str
    full_name: str  # Required now!
    phone: Optional[str] = None
    date_of_birth: Optional[str] = None
    occupation: Optional[str] = None
    monthly_budget: Optional[float] = None

class ExampleCategoryResponse(BaseModel):
    id: int
    name: str
    icon: str
    color: str
    description: Optional[str]
    sort_order: int
    
    class Config:
        from_attributes = True

class CategoryBatchCreate(BaseModel):
    category_ids: List[int]  # Example category IDs to add

class CategoryMigrateRequest(BaseModel):
    from_category_name: str
    to_category_name: str

class CategoryMigrateResponse(BaseModel):
    message: str
    affected_count: int
    from_category: str
    to_category: str


class ExampleCategory(Base):
    __tablename__ = "example_categories"
    __table_args__ = {"schema": "public"}  # 🔥 THIS FIXES IT

    id = Column(Integer, primary_key=True, index=True)
    name = Column(String(50), nullable=False, unique=True)
    icon = Column(String(50), nullable=False)
    color = Column(String(7), nullable=False)
    description = Column(String(200))
    sort_order = Column(Integer, default=0)

class CategoryMigration(Base):
    __tablename__ = "category_migrations"
    
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, ForeignKey("users.id"), nullable=False)
    from_category_name = Column(String(50), nullable=False)
    to_category_name = Column(String(50), nullable=False)
    affected_count = Column(Integer, nullable=False)
    migrated_at = Column(DateTime, default=datetime.utcnow)

# ==========================================
# APP INITIALIZATION
# ==========================================

app = FastAPI(title="Expense Tracker API", version="3.0.0")
app.include_router(sms_router)
app.include_router(voice_router)

@app.on_event("startup")
async def startup_event():
    print("=" * 60)
    print("🚀 EXPENSE TRACKER API STARTING UP")
    print("=" * 60)
    print(f"📍 Frontend URL: {FRONTEND_URL}")
    print(f"📍 Database: {DATABASE_URL[:40]}...")
    print(f"📧 Email enabled: {EMAIL_ENABLED}")
    print(f"🔑 Secret key configured: {bool(SECRET_KEY and SECRET_KEY != 'your-secret-key-change-in-production-PLEASE')}")
    print("=" * 60)

# CORS Configuration
allowed_origins = [
    "http://localhost:5173",
    "http://127.0.0.1:5173",
    "http://localhost:3000",
    "https://webapp-expense.vercel.app",
]

if FRONTEND_URL and FRONTEND_URL not in allowed_origins:
    allowed_origins.append(FRONTEND_URL)
    print(f"✅ Added FRONTEND_URL to CORS: {FRONTEND_URL}")

# Use allow_origin_regex to handle all Vercel preview URLs
app.add_middleware(
    CORSMiddleware,
    allow_origin_regex=r"https://.*\.vercel\.app",  # Allows all Vercel subdomains
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],
)

print(f"🔧 CORS configured with origins: {allowed_origins}")
print(f"🔧 Also allowing all *.vercel.app subdomains via regex")

# ==========================================
# DEPENDENCIES
# ==========================================

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# ==========================================
# HELPER FUNCTIONS
# ==========================================

def verify_password(plain_password: str, hashed_password: str) -> bool:
    sha256 = hashlib.sha256(plain_password.encode("utf-8")).hexdigest()
    return pwd_context.verify(sha256, hashed_password)

def get_password_hash(password: str) -> str:
    sha256 = hashlib.sha256(password.encode("utf-8")).hexdigest()
    return pwd_context.hash(sha256)

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def get_current_user(
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db)
):
    token = credentials.credentials
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        
        if username is None:
            raise HTTPException(status_code=401, detail="Invalid token")
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")
    
    user = db.query(User).filter(User.username == username).first()
    if user is None:
        raise HTTPException(status_code=401, detail="User not found")
    
    return user

def normalize(value: str | None) -> str | None:
    if not value:
        return None
    return value.strip().lower()

# ✅ NEW: Create Default Categories Helper
def create_default_categories(db: Session, user_id: int):
    """Create default categories for a new user"""
    default_categories = [
        {"name": "Food", "color": "#FF6B6B", "icon": "utensils"},
        {"name": "Transport", "color": "#4ECDC4", "icon": "car"},
        {"name": "Shopping", "color": "#95E1D3", "icon": "shopping-bag"},
        {"name": "Bills", "color": "#FFE66D", "icon": "file-text"},
        {"name": "Entertainment", "color": "#A8E6CF", "icon": "film"},
        {"name": "Income", "color": "#4CAF50", "icon": "indian-rupee"},
        {"name": "Education", "color": "#2196F3", "icon": "book"},
        {"name": "Health", "color": "#E91E63", "icon": "heart"},
    ]
    
    for cat_data in default_categories:
        category = Category(
            user_id=user_id,
            name=cat_data["name"],
            color=cat_data["color"],
            icon=cat_data["icon"]
        )
        db.add(category)
    
    db.commit()

def send_email(to_email: str, subject: str, html_body: str) -> bool:
    BREVO_API_KEY = os.getenv("BREVO_API_KEY")

    if not BREVO_API_KEY:
        print("❌ BREVO_API_KEY not set")
        return False

    url = "https://api.brevo.com/v3/smtp/email"

    headers = {
        "accept": "application/json",
        "content-type": "application/json",
        "api-key": BREVO_API_KEY,  # ✅ THIS IS REQUIRED
    }

    payload = {
        "sender": {
            "name": FROM_NAME,
            "email": FROM_EMAIL
        },
        "to": [
            {"email": to_email}
        ],
        "subject": subject,
        "htmlContent": html_body
    }

    try:
        res = requests.post(url, json=payload, headers=headers)

        if res.status_code >= 400:
            print("❌ Brevo error:", res.text)
            return False

        return True

    except Exception as e:
        print("❌ Email send failed:", str(e))
        return False
# ==========================================
# EXPORT/IMPORT HELPER FUNCTIONS
# ==========================================

# Check if openpyxl is available
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

def export_to_csv(expenses: List[Expense]) -> str:
    """Convert expenses to CSV format"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow(['Date', 'Amount', 'Category', 'Description', 'Type'])
    
    # Data
    for expense in expenses:
        writer.writerow([
            expense.date.strftime('%Y-%m-%d') if isinstance(expense.date, date_type) else expense.date,
            expense.amount,
            expense.category,
            expense.description,
            expense.type
        ])
    
    return output.getvalue()

def export_to_excel(expenses: List[Expense]) -> bytes:
    """Convert expenses to Excel format with formatting"""
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Excel support not available. Install openpyxl.")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses"
    
    # Header
    headers = ['Date', 'Amount', 'Category', 'Description', 'Type']
    ws.append(headers)
    
    # Style header
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for expense in expenses:
        ws.append([
            expense.date.strftime('%Y-%m-%d') if isinstance(expense.date, date_type) else expense.date,
            expense.amount,
            expense.category,
            expense.description,
            expense.type
        ])
    
    # Format amount column
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '$#,##0.00'
    
    # Auto-adjust column widths
    for i, column in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(i)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()

def parse_csv_file(file_content: bytes) -> List[dict]:
    """Parse CSV file and return list of expense dicts"""
    text_content = file_content.decode('utf-8')
    reader = csv.DictReader(io.StringIO(text_content))
    
    expenses = []
    for row in reader:
        # Validate required fields
        if not all(key in row for key in ['Date', 'Amount', 'Category', 'Description', 'Type']):
            raise HTTPException(status_code=400, detail="CSV must have columns: Date, Amount, Category, Description, Type")
        
        expenses.append({
            'date': row['Date'],
            'amount': float(row['Amount']),
            'category': row['Category'],
            'description': row['Description'],
            'type': row['Type'].lower()
        })
    
    return expenses

def normalize(name: str | None) -> str | None:
    if not name:
        return None
    return name.strip().lower()


def parse_excel_file(file_content: bytes) -> List[dict]:
    """Parse Excel file and return list of expense dicts"""
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=400, detail="Excel support not available")
    
    wb = openpyxl.load_workbook(io.BytesIO(file_content))
    ws = wb.active
    
    expenses = []
    headers = [cell.value for cell in ws[1]]
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        
        expenses.append({
            'date': str(row[0]),
            'amount': float(row[1]),
            'category': str(row[2]),
            'description': str(row[3]),
            'type': str(row[4]).lower()
        })
    
    return expenses

def get_splitwise_auth_header(user: User, db: Session) -> dict:
  if not user.splitwise_access_token:
      raise HTTPException(status_code=400, detail="Splitwise not connected")

  # refresh if expired
  if user.splitwise_token_expires_at:
      now = datetime.utcnow()
      # Handle both timezone-aware and naive datetimes
      expires_at = user.splitwise_token_expires_at
      if hasattr(expires_at, 'tzinfo') and expires_at.tzinfo is not None:
          # expires_at is timezone-aware, make now timezone-aware too
          from datetime import timezone
          now = datetime.now(timezone.utc)
      
      if expires_at < now:
          token_url = f"{SPLITWISE_BASE_URL}/oauth/token"
          data = {
              "grant_type": "refresh_token",
              "refresh_token": user.splitwise_refresh_token,
              "client_id": SPLITWISE_CLIENT_ID,
              "client_secret": SPLITWISE_CLIENT_SECRET,
          }
          resp = requests.post(token_url, data=data)
          if resp.status_code != 200:
              raise HTTPException(status_code=400, detail="Failed to refresh Splitwise token")

          tokens = resp.json()
          user.splitwise_access_token = tokens["access_token"]
          user.splitwise_refresh_token = tokens.get("refresh_token", user.splitwise_refresh_token)
          expires_in = tokens.get("expires_in", 3600)
          user.splitwise_token_expires_at = datetime.utcnow() + timedelta(seconds=expires_in)
          db.commit()

  return {"Authorization": f"Bearer {user.splitwise_access_token}"}



def categorize_with_ai(description: str, user: User, db: Session) -> str:
    """
    Use Claude AI to categorize expense based on description and user's categories
    """
    ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
    
    # Get user's categories
    user_categories = db.query(Category).filter(
        Category.user_id == user.id
    ).all()
    
    category_names = [cat.name for cat in user_categories]
    
    # If no categories or no API key, use fallback
    if not category_names or not ANTHROPIC_API_KEY:
        return map_keywords(description)
    
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        
        prompt = f"""You are categorizing an expense. Choose the BEST matching category from the user's list.

Expense Description: "{description}"

User's Categories:
{', '.join(category_names)}

Instructions:
- Return ONLY the category name, nothing else
- Choose the most appropriate category from the list above
- If nothing matches well, use the most general category available
- DO NOT return a category not in the list

Category:"""

        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=50,
            messages=[{"role": "user", "content": prompt}]
        )
        
        category = message.content[0].text.strip()
        
        # Validate it's in user's categories
        if category in category_names:
            return category
        
        # If AI returned invalid category, use fallback
        return map_keywords(description)
        
    except Exception as e:
        print(f"AI categorization failed: {e}")
        return map_keywords(description)


def map_keywords(text: str) -> str:
    """Fallback keyword-based categorization"""
    t = (text or "").lower()
    if any(k in t for k in ["food", "restaurant", "dinner", "lunch", "pizza", "coffee"]):
        return "Food"
    if any(k in t for k in ["uber", "ola", "taxi", "train", "flight", "bus", "travel"]):
        return "Transport"
    if any(k in t for k in ["rent", "electricity", "wifi", "bill"]):
        return "Bills"
    if any(k in t for k in ["amazon", "shopping", "clothes", "shirt", "jeans"]):
        return "Shopping"
    return "Other"

def sync_splitwise_for_user(db: Session, user: User, mode: str = "today") -> int:
    if not user.splitwise_access_token:
        return 0

    headers = get_splitwise_auth_header(user, db)

    params = {}
    if mode == "today":
        start_of_today = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        params["dated_after"] = start_of_today.isoformat()

    resp = requests.get(
        f"{SPLITWISE_BASE_URL}/api/v3.0/get_expenses",
        params=params,
        headers=headers,
    )
    if resp.status_code != 200:
        print("Splitwise error:", resp.text)
        return 0

    data = resp.json()
    expenses = data.get("expenses", [])
    imported = 0

    for sw in expenses:
        sw_id = sw.get("id")
        if sw_id is None:
            continue

        # Avoid duplicates
        existing = db.query(PendingTransaction).filter(
            PendingTransaction.user_id == user.id,
            PendingTransaction.splitwise_expense_id == sw_id,
        ).first()
        if existing:
            continue

        # Find this user’s record in the expense
        me = None
        for u in sw.get("users", []):
            if u.get("user_id") == user.splitwise_user_id:
                me = u
                break

        if not me:
            continue

        owed_raw = me.get("owed_share") or "0"
        try:
            owed = float(owed_raw)
        except ValueError:
            owed = 0.0

        if owed <= 0:
            continue

        description = sw.get("description") or "Splitwise Expense"
        sw_date = sw.get("date") or datetime.utcnow().isoformat()
        date_str = sw_date.split("T")[0]

        category = categorize_with_ai(description, user, db)

        pending = PendingTransaction(
            user_id=user.id,
            token=token_urlsafe(16),
            amount=owed,
            description=description,
            category=category,
            date=date_str,
            type="expense",
            status="pending",
            splitwise_expense_id=sw_id,
            splitwise_group_name=(sw.get("group") or {}).get("name"),
            splitwise_raw_json=json.dumps(sw),
        )

        db.add(pending)
        imported += 1

    user.splitwise_last_sync_at = datetime.utcnow()
    db.commit()
    return imported

# ==========================================
# AUTH ROUTES
# ==========================================

@app.post("/api/splitwise/sync-today")
def sync_today(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    imported = sync_splitwise_for_user(db, current_user, mode="today")
    return {"imported": imported}

@app.post("/api/splitwise/sync-all-user")
def sync_all_user(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    imported = sync_splitwise_for_user(db, current_user, mode="all")
    return {"imported": imported}

@app.api_route("/api/splitwise/sync-all", methods=["GET", "POST"])
def sync_all(secret: str, db: Session = Depends(get_db)):
    if secret != CRON_SECRET:
        raise HTTPException(status_code=401, detail="Unauthorized")

    users = db.query(User).filter(User.splitwise_access_token.isnot(None)).all()
    total = 0
    for user in users:
        try:
            total += sync_splitwise_for_user(db, user, mode="today")
        except Exception as e:
            print(f"Sync failed for user {user.id}: {e}")

    return {"users": len(users), "imported": total}


@app.post("/api/splitwise/sync")
def trigger_splitwise_sync(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    count = sync_splitwise_for_user(db, current_user)
    return {"imported": count}

CRON_SECRET = os.getenv("CRON_SECRET", "change-me")

@app.post("/api/splitwise/sync-all")
def sync_splitwise_all(
    secret: str,
    db: Session = Depends(get_db)
):
    if secret != CRON_SECRET:
        raise HTTPException(status_code=401, detail="Unauthorized")

    users = db.query(User).filter(User.splitwise_access_token.isnot(None)).all()
    total = 0
    for user in users:
        total += sync_splitwise_for_user(db, user)
    return {"total_imported": total, "users": len(users)}


from urllib.parse import urlencode

@app.get("/api/splitwise/auth-url")
def get_splitwise_auth_url(current_user: User = Depends(get_current_user)):
    params = {
        "response_type": "code",
        "client_id": SPLITWISE_CLIENT_ID,
        "redirect_uri": SPLITWISE_REDIRECT_URI,
        "state": current_user.username,
    }
    url = f"{SPLITWISE_BASE_URL}/oauth/authorize?{urlencode(params)}"
    return {"auth_url": url}

@app.get("/api/splitwise/debug-config")
def debug_splitwise_config(current_user: User = Depends(get_current_user)):
    """Debug endpoint to check Splitwise configuration"""
    return {
        "client_id_present": bool(SPLITWISE_CLIENT_ID),
        "client_secret_present": bool(SPLITWISE_CLIENT_SECRET),
        "redirect_uri": SPLITWISE_REDIRECT_URI,
        "frontend_url": FRONTEND_URL,
        "api_base": API_BASE,
    }

@app.get("/api/splitwise/callback")
def splitwise_callback(code: str, state: str, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == state).first()
    if not user:
        raise HTTPException(status_code=400, detail="Invalid state")

    token_url = f"{SPLITWISE_BASE_URL}/oauth/token"
    data = {
        "grant_type": "authorization_code",
        "code": code,
        "client_id": SPLITWISE_CLIENT_ID,
        "client_secret": SPLITWISE_CLIENT_SECRET,
        "redirect_uri": SPLITWISE_REDIRECT_URI,
    }
    
    print(f"🔹 Exchanging token with Splitwise...")
    print(f"   Redirect URI: {SPLITWISE_REDIRECT_URI}")
    print(f"   Client ID present: {bool(SPLITWISE_CLIENT_ID)}")
    print(f"   Client Secret present: {bool(SPLITWISE_CLIENT_SECRET)}")
    
    resp = requests.post(token_url, data=data)
    if resp.status_code != 200:
        error_detail = resp.text
        print(f"❌ Splitwise token exchange failed!")
        print(f"   Status: {resp.status_code}")
        print(f"   Error: {error_detail}")
        
        # Return more helpful error
        if "redirect_uri" in error_detail.lower():
            raise HTTPException(
                status_code=500, 
                detail=f"Redirect URI mismatch. Expected: {SPLITWISE_REDIRECT_URI}"
            )
        elif "client" in error_detail.lower():
            raise HTTPException(
                status_code=500,
                detail="Invalid Client ID or Secret. Check environment variables."
            )
        else:
            raise HTTPException(
                status_code=500, 
                detail=f"Splitwise OAuth failed: {error_detail[:200]}"
            )

    tokens = resp.json()
    user.splitwise_access_token = tokens["access_token"]
    user.splitwise_refresh_token = tokens.get("refresh_token")
    expires_in = tokens.get("expires_in", 3600)
    user.splitwise_token_expires_at = datetime.utcnow() + timedelta(seconds=expires_in)

    # Get current user from Splitwise API so we know their Splitwise user id
    headers = {"Authorization": f"Bearer {user.splitwise_access_token}"}
    me_resp = requests.get(
        f"{SPLITWISE_BASE_URL}/api/v3.0/get_current_user",
        headers=headers,
    )
    if me_resp.status_code == 200:
        me_data = me_resp.json()
        sw_user = me_data.get("user") or {}
        user.splitwise_user_id = sw_user.get("id")
        print(f"✅ Splitwise User ID saved: {user.splitwise_user_id}")
    else:
        print(f"⚠️  Failed to get Splitwise user ID: {me_resp.status_code}")

    db.commit()
    print(f"✅ Tokens saved to database for user: {user.username}")

    # Auto-sync today's expenses once after connect
    try:
        sync_splitwise_for_user(db, user, mode="today")
    except Exception as e:
        print("Initial sync failed:", e)

    # Redirect back to frontend with success indicator
    final_url = f"{FRONTEND_URL}/splitwise?connected=true"
    return RedirectResponse(url=final_url)


@app.post("/api/signup", response_model=Token)
def signup(user: SignupWithProfile, db: Session = Depends(get_db)):
    """
    Enhanced signup with profile information
    full_name is now REQUIRED
    """
    print(f"\n🔵 SIGNUP REQUEST RECEIVED")
    print(f"   Username: {user.username}")
    print(f"   Email: {user.email}")
    print(f"   Full name: {user.full_name}")
    
    try:
        # Check existing user
        if db.query(User).filter(User.username == user.username).first():
            print(f"❌ Username already exists: {user.username}")
            raise HTTPException(status_code=400, detail="Username already registered")
        if db.query(User).filter(User.email == user.email).first():
            print(f"❌ Email already exists: {user.email}")
            raise HTTPException(status_code=400, detail="Email already registered")
        
        # Validate full_name is provided
        if not user.full_name or len(user.full_name.strip()) == 0:
            print(f"❌ Full name is empty or invalid")
            raise HTTPException(status_code=400, detail="Full name is required")
        
        print(f"✅ Validation passed, creating user...")
        
        # Create user with profile data
        hashed_password = get_password_hash(user.password)
        new_user = User(
            username=user.username,
            email=user.email,
            hashed_password=hashed_password,
            full_name=user.full_name.strip(),
            phone=user.phone,
            date_of_birth=datetime.strptime(user.date_of_birth, "%Y-%m-%d").date() if user.date_of_birth else None,
            occupation=user.occupation,
            monthly_budget=user.monthly_budget,
            onboarding_completed=False  # Will be set to True after category selection
        )
        
        db.add(new_user)
        db.commit()
        db.refresh(new_user)
        
        print(f"✅ User created successfully: {new_user.username} (ID: {new_user.id})")
        
        # NO AUTO-CATEGORY CREATION - User will select during onboarding
        
        # Create token
        access_token = create_access_token(data={"sub": new_user.username})
        
        print(f"✅ Token generated for: {new_user.username}\n")
        return {"access_token": access_token, "token_type": "bearer"}
        
    except HTTPException as e:
        print(f"❌ HTTP Exception: {e.status_code} - {e.detail}\n")
        raise
    except Exception as e:
        print(f"❌ Unexpected error in signup: {str(e)}")
        import traceback
        traceback.print_exc()
        print()
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@app.post("/api/login", response_model=Token)
def login(user: UserLogin, db: Session = Depends(get_db)):
    db_user = db.query(User).filter(User.username == user.username).first()
    if not db_user or not verify_password(user.password, db_user.hashed_password):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    access_token = create_access_token(data={"sub": db_user.username})
    return {"access_token": access_token, "token_type": "bearer"}

@app.post("/api/forgot-password")
def forgot_password(request: ForgotPasswordRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.email == request.email).first()
    
    if not user:
        raise HTTPException(status_code=404, detail="Email not found")
    
    # Generate reset token
    reset_token = token_urlsafe(32)
    expires_at = datetime.utcnow() + timedelta(hours=1)
    
    token_entry = PasswordResetToken(
        user_id=user.id,
        token=reset_token,
        expires_at=expires_at
    )
    db.add(token_entry)
    db.commit()
    
    # ✅ UPDATED: Create reset URL with token in query parameter
    reset_url = f"{FRONTEND_URL}/reset-password?token={reset_token}"
    
    # Send email
    subject = "🔐 Reset Your Password - Expense Tracker"
    html_body = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
            .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
            .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px; text-align: center; border-radius: 10px 10px 0 0; }}
            .content {{ background: #f9f9f9; padding: 30px; border-radius: 0 0 10px 10px; }}
            .button {{ display: inline-block; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 15px 30px; text-decoration: none; border-radius: 5px; margin: 20px 0; font-weight: bold; }}
            .link-box {{ background: white; padding: 15px; border: 1px solid #ddd; border-radius: 5px; word-break: break-all; margin: 15px 0; font-family: monospace; font-size: 12px; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>🔐 Password Reset</h1>
            </div>
            <div class="content">
                <p>Hi {user.username},</p>
                <p>You requested to reset your password. Click the button below to reset it:</p>
                
                <div style="text-align: center;">
                    <a href="{reset_url}" class="button">Reset Password Now →</a>
                </div>
                
                <p>Or copy and paste this link into your browser:</p>
                <div class="link-box">{reset_url}</div>
                
                <p><strong>This link expires in 1 hour.</strong></p>
                <p>If you didn't request this, you can safely ignore this email.</p>
                
                <hr style="margin: 30px 0; border: none; border-top: 1px solid #ddd;">
                <p style="color: #666; font-size: 12px;">Expense Tracker - Your Personal Finance Manager</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    email_sent = send_email(user.email, subject, html_body)
    
    # ✅ UPDATED: Return both reset_url and token in dev mode
    if os.getenv("ENVIRONMENT") == "development":
        return {
            "message": "Password reset email sent" if email_sent else "Email not configured (dev mode)",
            "reset_url": reset_url,  # ✅ Added for dev mode
            "token": reset_token
        }
    
    return {"message": "Password reset email sent" if email_sent else "Email not configured"}

@app.post("/api/reset-password")
def reset_password(request: ResetPasswordRequest, db: Session = Depends(get_db)):
    token_entry = db.query(PasswordResetToken).filter(
        PasswordResetToken.token == request.token,
        PasswordResetToken.used == 0
    ).first()
    
    if not token_entry:
        raise HTTPException(status_code=400, detail="Invalid or expired token")
    
    if token_entry.expires_at < datetime.utcnow():
        raise HTTPException(status_code=400, detail="Token has expired")
    
    user = db.query(User).filter(User.id == token_entry.user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Update password
    user.hashed_password = get_password_hash(request.new_password)
    token_entry.used = 1
    db.commit()
    
    return {"message": "Password reset successful"}

# ==========================================
# iOS SHORTCUT SMS PARSER ENDPOINT
# ==========================================

@app.get("/api/user/sms-parse")
async def user_sms_parse(
    sms: str = Query(..., description="SMS message text"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    User-specific SMS parser for iOS Shortcuts
    Requires authentication token in Authorization header
    
    Usage from iOS Shortcut:
    GET /api/user/sms-parse?sms=Your%20A/c%20debited%20Rs.500...
    Headers: Authorization: Bearer {user_token}
    
    Returns:
    {
        "success": true,
        "url": "https://webapp-expense.vercel.app/add-expense/{token}",
        "parsed_data": {...},
        "already_processed": false
    }
    """
    from sms_parser_api import parse_sms_with_claude
    import secrets
    
    # Create hash of SMS to check for duplicates
    sms_hash = hashlib.sha256(sms.encode()).hexdigest()
    
    # Check if this SMS has already been processed in last 24 hours
    one_day_ago = datetime.now() - timedelta(days=1)
    existing = db.query(PendingTransaction).filter(
        PendingTransaction.user_id == current_user.id,
        PendingTransaction.description.contains(sms_hash[:16]),  # Store partial hash in description
        PendingTransaction.created_at >= one_day_ago
    ).first()
    
    if existing:
        # SMS already processed - return redirect to dashboard
        return {
            "success": True,
            "url": f"{FRONTEND_URL}/",  # Redirect to dashboard
            "already_processed": True,
            "message": "This SMS was already processed"
        }
    
    # Parse SMS using Claude AI
    parse_result = parse_sms_with_claude(sms)
    
    if not parse_result.get("success"):
        raise HTTPException(status_code=400, detail="Failed to parse SMS")
    
    data = parse_result["data"]
    
    # Create description with hash for deduplication
    original_description = data.get("merchant", "Unknown")
    description_with_hash = f"{original_description} [#{sms_hash[:16]}]"
    
    # Create pending transaction
    pending = PendingTransaction(
        user_id=current_user.id,
        amount=data.get("amount"),
        description=description_with_hash,  # Include hash for dedup
        category=data.get("category", "Other"),
        date=data.get("date", datetime.now().strftime("%Y-%m-%d")),
        type="income" if data.get("transaction_type") == "credit" else "expense",
        token=secrets.token_urlsafe(16),
        status="pending"
    )
    db.add(pending)
    db.commit()
    
    # Return URL for the shortcut to open
    expense_url = f"{FRONTEND_URL}/add-expense/{pending.token}"
    
    return {
        "success": True,
        "url": expense_url,
        "parsed_data": data,
        "confidence": data.get("confidence", 0.5),
        "already_processed": False
    }

# ==========================================
# GENERATE PERSONALIZED SHORTCUT URL
# ==========================================

@app.post("/api/user/regenerate-shortcut-url")
async def regenerate_shortcut_url(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Generate a fresh shortcut URL
    Note: This doesn't actually invalidate old URLs (they expire in 30 days)
    Just generates a new token with same validity
    """
    # Generate new token
    shortcut_token_data = {"sub": current_user.username}
    shortcut_expire = datetime.utcnow() + timedelta(days=30)
    shortcut_token_data.update({"exp": shortcut_expire})
    shortcut_token = jwt.encode(shortcut_token_data, SECRET_KEY, algorithm=ALGORITHM)
    
    # URL encode the token
    encoded_token = quote(shortcut_token, safe='')
    
    personalized_url = f"{FRONTEND_URL}/add-expense-from-sms?token={encoded_token}&sms"
    
    return {
        "success": True,
        "shortcut_url": personalized_url,
        "message": "New shortcut URL generated. Update your iOS Shortcut with this URL.",
        "expires_in_days": 30
    }

@app.get("/api/user/shortcut-url")
async def get_personalized_shortcut_url(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Generate a personalized URL for iOS Shortcuts that includes the user's auth token
    This URL can be directly used in iOS Shortcuts without manual token configuration
    
    Returns a complete URL that:
    1. Accepts SMS text as query parameter
    2. Has auth token embedded
    3. Returns the expense confirmation URL
    
    Usage: User copies this URL and pastes it in their iOS Shortcut
    """
    # Create a long-lived token (30 days) specifically for shortcuts
    shortcut_token_data = {"sub": current_user.username}
    shortcut_expire = datetime.utcnow() + timedelta(days=30)
    shortcut_token_data.update({"exp": shortcut_expire})
    shortcut_token = jwt.encode(shortcut_token_data, SECRET_KEY, algorithm=ALGORITHM)
    
    # URL encode the token to handle special characters (dots in JWT)
    encoded_token = quote(shortcut_token, safe='')
    
    # Generate the personalized FRONTEND URL with embedded token
    # User pastes this in iOS Shortcut, it goes directly to frontend with auth
    personalized_url = f"{FRONTEND_URL}/add-expense-from-sms?token={encoded_token}&sms"
    
    return {
        "success": True,
        "shortcut_url": personalized_url,
        "instructions": "Paste this URL in your iOS Shortcut. Replace {SMS_TEXT} with the SMS text variable.",
        "example": f"{FRONTEND_URL}/add-expense-from-sms?token={encoded_token}&sms=Your%20A/c%20debited%20Rs.500%20at%20Starbucks",
        "expires_in_days": 30,
        "note": "This URL works even without logging in - the token is embedded"
    }

@app.get("/api/user/sms-parse-public")
async def user_sms_parse_public(
    token: str = Query(..., description="User's shortcut token"),
    sms: str = Query(..., description="SMS message text"),
    db: Session = Depends(get_db)
):
    """
    Public SMS parser endpoint that accepts token as query parameter
    Used by iOS Shortcuts with personalized URL
    
    This endpoint doesn't require Authorization header - token is in the URL
    """
    import secrets
    
    # Verify token
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise HTTPException(status_code=401, detail="Invalid token")
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    
    # Get user
    user = db.query(User).filter(User.username == username).first()
    if user is None:
        raise HTTPException(status_code=401, detail="User not found")
    
    # Parse SMS using Claude AI
    from sms_parser_api import parse_sms_with_claude
    parse_result = parse_sms_with_claude(sms)
    
    if not parse_result.get("success"):
        raise HTTPException(status_code=400, detail="Failed to parse SMS")
    
    data = parse_result["data"]
    
    # Create pending transaction
    pending = PendingTransaction(
        user_id=user.id,
        amount=data.get("amount"),
        description=data.get("merchant", "Unknown"),
        category=data.get("category", "Other"),
        date=data.get("date", datetime.now().strftime("%Y-%m-%d")),
        type="income" if data.get("transaction_type") == "credit" else "expense",
        token=secrets.token_urlsafe(16),
        status="pending"
    )
    db.add(pending)
    db.commit()
    
    # Return URL for the shortcut to open
    expense_url = f"{FRONTEND_URL}/add-expense/{pending.token}"
    
    return {
        "success": True,
        "url": expense_url,
        "parsed_data": data,
        "confidence": data.get("confidence", 0.5)
    }

# ==========================================
# CATEGORY ROUTES
# ==========================================

@app.get("/api/categories", response_model=List[CategoryResponse])
def get_categories(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get all categories for current user"""
    categories = db.query(Category).filter(
        Category.user_id == current_user.id
    ).order_by(Category.name).all()
    
    # If user has no categories, create defaults
    if not categories:
        create_default_categories(db, current_user.id)
        categories = db.query(Category).filter(
            Category.user_id == current_user.id
        ).order_by(Category.name).all()
    
    return categories

@app.post("/api/categories", response_model=CategoryResponse)
def create_category(
    category: CategoryCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Create a new category"""
    # Check if category name already exists for this user
    existing = db.query(Category).filter(
        Category.user_id == current_user.id,
        Category.name == category.name
    ).first()
    
    if existing:
        raise HTTPException(
            status_code=400,
            detail=f"Category '{category.name}' already exists"
        )
    
    # Validate color format
    if category.color and not category.color.startswith("#"):
        raise HTTPException(
            status_code=400,
            detail="Color must be in hex format (e.g., #FF5733)"
        )
    
    new_category = Category(
        user_id=current_user.id,
        name=category.name,
        color=category.color,
        icon=category.icon
    )
    
    db.add(new_category)
    db.commit()
    db.refresh(new_category)
    
    return new_category

@app.put("/api/categories/{category_id}", response_model=CategoryResponse)
def update_category(
    category_id: int,
    category_update: CategoryUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update a category"""
    category = db.query(Category).filter(
        Category.id == category_id,
        Category.user_id == current_user.id
    ).first()
    
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    
    # Check if new name conflicts
    if category_update.name and category_update.name != category.name:
        existing = db.query(Category).filter(
            Category.user_id == current_user.id,
            Category.name == category_update.name
        ).first()
        
        if existing:
            raise HTTPException(
                status_code=400,
                detail=f"Category '{category_update.name}' already exists"
            )
        
        category.name = category_update.name
    
    if category_update.color:
        if not category_update.color.startswith("#"):
            raise HTTPException(
                status_code=400,
                detail="Color must be in hex format (e.g., #FF5733)"
            )
        category.color = category_update.color
    
    if category_update.icon:
        category.icon = category_update.icon
    
    db.commit()
    db.refresh(category)
    
    return category

@app.delete("/api/categories/{category_id}")
def delete_category(
    category_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Delete a category"""
    category = db.query(Category).filter(
        Category.id == category_id,
        Category.user_id == current_user.id
    ).first()
    
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    
    # Prevent deletion of Income category
    if category.name == "Income":
        raise HTTPException(
            status_code=400,
            detail="Income category cannot be deleted as it's essential for tracking income"
        )
    
    # Check if category is used
    expense_count = db.query(Expense).filter(
        Expense.user_id == current_user.id,
        Expense.category == category.name
    ).count()
    
    if expense_count > 0:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete category '{category.name}' - it's used in {expense_count} expense(s)"
        )
    
    db.delete(category)
    db.commit()
    
    return {"message": f"Category '{category.name}' deleted successfully"}

@app.get("/api/categories/stats")
def get_category_stats(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    # 1. Aggregate directly from expenses (SOURCE OF TRUTH)
    expense_stats = (
        db.query(
            Expense.category.label("category"),
            func.count(Expense.id).label("expense_count"),
            func.sum(Expense.amount).label("total_amount"),
        )
        .filter(
            Expense.user_id == current_user.id,
            Expense.type == "expense"
        )
        .group_by(Expense.category)
        .all()
    )

    # 2. Fetch user categories for enrichment (OPTIONAL)
    categories = {
        c.name: c
        for c in db.query(Category)
        .filter(Category.user_id == current_user.id)
        .all()
    }

    # 3. Merge safely
    stats = []
    for category, count, total in expense_stats:
        cat = categories.get(normalize(category))

        stats.append({
            "category": category,
            "color": cat.color if cat else "#999999",
            "icon": cat.icon if cat else "📦",
            "expense_count": count,
            "total_amount": float(total),
            "can_delete": count == 0 and category != "Income",
        })

    return stats

# ==========================================
# EXPENSE ROUTES
# ==========================================

@app.get("/api/expenses", response_model=List[ExpenseResponse])
def get_expenses(
    category: Optional[str] = None,
    type: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    query = db.query(Expense).filter(Expense.user_id == current_user.id)
    
    if category:
        query = query.filter(Expense.category == category)
    if type:
        query = query.filter(Expense.type == type)
    
    expenses = query.order_by(Expense.date.desc()).all()
    
    return [
        ExpenseResponse(
            id=e.id,
            user_id=e.user_id,
            amount=e.amount,
            category=e.category,
            description=e.description,
            date=e.date.isoformat(),
            type=e.type
        ) for e in expenses
    ]

@app.post("/api/expenses", response_model=ExpenseResponse)
def create_expense(
    expense: ExpenseCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    new_expense = Expense(
        user_id=current_user.id,
        amount=expense.amount,
        category=expense.category,
        description=expense.description,
        date=datetime.strptime(expense.date, "%Y-%m-%d").date(),
        type=expense.type
    )
    db.add(new_expense)
    db.commit()
    db.refresh(new_expense)
    
    return ExpenseResponse(
        id=new_expense.id,
        user_id=new_expense.user_id,
        amount=new_expense.amount,
        category=new_expense.category,
        description=new_expense.description,
        date=new_expense.date.isoformat(),
        type=new_expense.type
    )

@app.put("/api/expenses/{expense_id}", response_model=ExpenseResponse)
def update_expense(
    expense_id: int,
    expense: ExpenseUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    db_expense = db.query(Expense).filter(
        Expense.id == expense_id,
        Expense.user_id == current_user.id
    ).first()
    
    if not db_expense:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    if expense.amount is not None:
        db_expense.amount = expense.amount
    if expense.category is not None:
        db_expense.category = expense.category
    if expense.description is not None:
        db_expense.description = expense.description
    if expense.date is not None:
        db_expense.date = datetime.strptime(expense.date, "%Y-%m-%d").date()
    if expense.type is not None:
        db_expense.type = expense.type
    
    db.commit()
    db.refresh(db_expense)
    
    return ExpenseResponse(
        id=db_expense.id,
        user_id=db_expense.user_id,
        amount=db_expense.amount,
        category=db_expense.category,
        description=db_expense.description,
        date=db_expense.date.isoformat(),
        type=db_expense.type
    )

@app.delete("/api/expenses/{expense_id}")
def delete_expense(
    expense_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    db_expense = db.query(Expense).filter(
        Expense.id == expense_id,
        Expense.user_id == current_user.id
    ).first()
    
    if not db_expense:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    db.delete(db_expense)
    db.commit()
    
    return {"message": "Expense deleted"}

# ==========================================
# EXPORT/IMPORT ROUTES
# ==========================================

@app.get("/api/export/csv")
def export_expenses_csv(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    type: Optional[str] = Query(None),
    email: Optional[bool] = Query(False),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export expenses as CSV"""
    query = db.query(Expense).filter(Expense.user_id == current_user.id)
    
    if start_date:
        query = query.filter(Expense.date >= datetime.strptime(start_date, "%Y-%m-%d").date())
    if end_date:
        query = query.filter(Expense.date <= datetime.strptime(end_date, "%Y-%m-%d").date())
    if category:
        query = query.filter(Expense.category == category)
    if type:
        query = query.filter(Expense.type == type)
    
    expenses = query.order_by(Expense.date.desc()).all()
    
    if not expenses:
        raise HTTPException(status_code=404, detail="No expenses found")
    
    csv_content = export_to_csv(expenses)
    
    if email:
        subject = "📊 Your Expense Export"
        html_body = f"""
        <html>
        <body>
            <h2>Your Expense Export</h2>
            <p>Hi {current_user.username},</p>
            <p>Your expense export is attached.</p>
            <p><strong>{len(expenses)} expenses</strong> exported.</p>
        </body>
        </html>
        """
        # TODO: Implement email with attachment
        return {"message": "Email sent successfully", "count": len(expenses)}
    
    filename = f"expenses_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    return StreamingResponse(
        iter([csv_content]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/api/export/excel")
def export_expenses_excel(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    type: Optional[str] = Query(None),
    email: Optional[bool] = Query(False),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export expenses as Excel"""
    query = db.query(Expense).filter(Expense.user_id == current_user.id)
    
    if start_date:
        query = query.filter(Expense.date >= datetime.strptime(start_date, "%Y-%m-%d").date())
    if end_date:
        query = query.filter(Expense.date <= datetime.strptime(end_date, "%Y-%m-%d").date())
    if category:
        query = query.filter(Expense.category == category)
    if type:
        query = query.filter(Expense.type == type)
    
    expenses = query.order_by(Expense.date.desc()).all()
    
    if not expenses:
        raise HTTPException(status_code=404, detail="No expenses found")
    
    excel_content = export_to_excel(expenses)
    
    filename = f"expenses_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        io.BytesIO(excel_content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.post("/api/import")
async def import_expenses(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Import expenses from CSV or Excel file"""
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")
    
    file_ext = file.filename.split('.')[-1].lower()
    
    if file_ext not in ['csv', 'xlsx', 'xls']:
        raise HTTPException(status_code=400, detail="File must be CSV or Excel")
    
    content = await file.read()
    
    # Parse file
    try:
        if file_ext == 'csv':
            expenses_data = parse_csv_file(content)
        else:
            expenses_data = parse_excel_file(content)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")
    
    # Import expenses
    imported = 0
    failed = 0
    errors = []
    
    for idx, exp_data in enumerate(expenses_data, 1):
        try:
            new_expense = Expense(
                user_id=current_user.id,
                amount=exp_data['amount'],
                category=exp_data['category'],
                description=exp_data['description'],
                date=datetime.strptime(exp_data['date'], "%Y-%m-%d").date(),
                type=exp_data['type']
            )
            db.add(new_expense)
            imported += 1
        except Exception as e:
            failed += 1
            errors.append(f"Row {idx}: {str(e)}")
    
    db.commit()
    
    return {
        "message": "Import completed",
        "imported": imported,
        "failed": failed,
        "total": len(expenses_data),
        "errors": errors[:10]  # Return first 10 errors
    }

# ==========================================
# PENDING TRANSACTION ROUTES
# ==========================================

@app.post("/api/generate-payment-url")
def generate_payment_url(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    unique_token = token_urlsafe(16)
    
    pending = PendingTransaction(
        user_id=current_user.id,
        token=unique_token,
        status="pending"
    )
    db.add(pending)
    db.commit()
    
    url = f"{FRONTEND_URL}/add-expense/{unique_token}"
    return {"url": url, "token": unique_token}

@app.get("/api/pending-transactions", response_model=List[PendingTransactionResponse])
def get_pending_transactions(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    pending = db.query(PendingTransaction).filter(
        PendingTransaction.user_id == current_user.id,
        PendingTransaction.status == "pending"
    ).order_by(PendingTransaction.created_at.desc()).all()
    
    return pending

@app.get("/api/pending-transaction/{token}")
def get_pending_transaction(token: str, db: Session = Depends(get_db)):
    pending = db.query(PendingTransaction).filter(
        PendingTransaction.token == token,
        PendingTransaction.status == "pending"
    ).first()
    
    if not pending:
        raise HTTPException(status_code=404, detail="Invalid or expired token")
    
    return {
        "token": pending.token,
        "amount": pending.amount,
        "category": pending.category,
        "description": pending.description,
        "date": pending.date,
        "type": pending.type
    }

@app.put("/api/pending-transaction/{token}")
def update_pending_transaction(
    token: str,
    data: PendingTransactionCreate,
    db: Session = Depends(get_db)
):
    pending = db.query(PendingTransaction).filter(
        PendingTransaction.token == token,
        PendingTransaction.status == "pending"
    ).first()
    
    if not pending:
        raise HTTPException(status_code=404, detail="Invalid or expired token")
    
    if data.amount is not None:
        pending.amount = data.amount
    if data.category is not None:
        pending.category = data.category
    if data.description is not None:
        pending.description = data.description
    if data.date is not None:
        pending.date = data.date
    if data.type is not None:
        pending.type = data.type
    
    db.commit()
    return {"message": "Updated"}

@app.post("/api/pending-transaction/{token}/approve")
def approve_pending_transaction(token: str, db: Session = Depends(get_db)):
    pending = db.query(PendingTransaction).filter(
        PendingTransaction.token == token,
        PendingTransaction.status == "pending"
    ).first()
    
    if not pending:
        raise HTTPException(status_code=404, detail="Invalid or expired token")
    
    if not all([pending.amount, pending.category, pending.description, pending.date, pending.type]):
        raise HTTPException(status_code=400, detail="Incomplete transaction data")
    
    new_expense = Expense(
        user_id=pending.user_id,
        amount=pending.amount,
        category=pending.category,
        description=pending.description,
        date=datetime.strptime(pending.date, "%Y-%m-%d").date(),
        type=pending.type
    )
    db.add(new_expense)
    
    pending.status = "approved"
    db.commit()
    
    return {"message": "Transaction approved"}

@app.delete("/api/pending-transaction/{token}")
def delete_pending_transaction(token: str, db: Session = Depends(get_db)):
    pending = db.query(PendingTransaction).filter(
        PendingTransaction.token == token
    ).first()
    
    if not pending:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    db.delete(pending)
    db.commit()
    
    return {"message": "Transaction deleted"}


# ==========================================
# PROFILE ROUTES
# ==========================================

@app.get("/api/profile", response_model=UserProfileResponse)
def get_profile(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get current user's profile"""
    return UserProfileResponse(
        username=current_user.username,
        email=current_user.email,
        full_name=current_user.full_name,
        phone=current_user.phone,
        date_of_birth=current_user.date_of_birth.isoformat() if current_user.date_of_birth else None,
        occupation=current_user.occupation,
        monthly_budget=current_user.monthly_budget,
        onboarding_completed=current_user.onboarding_completed or False,
        created_at=current_user.created_at if hasattr(current_user, 'created_at') else None,
        splitwise_user_id=current_user.splitwise_user_id,
        splitwise_last_sync_at=current_user.splitwise_last_sync_at
    )

@app.put("/api/profile", response_model=UserProfileResponse)
def update_profile(
    profile_data: UserProfile,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update user profile"""
    if profile_data.full_name is not None:
        current_user.full_name = profile_data.full_name
    if profile_data.phone is not None:
        current_user.phone = profile_data.phone
    if profile_data.date_of_birth is not None:
        current_user.date_of_birth = datetime.strptime(profile_data.date_of_birth, "%Y-%m-%d").date()
    if profile_data.occupation is not None:
        current_user.occupation = profile_data.occupation
    if profile_data.monthly_budget is not None:
        current_user.monthly_budget = profile_data.monthly_budget
    
    db.commit()
    db.refresh(current_user)
    
    return UserProfileResponse(
        username=current_user.username,
        email=current_user.email,
        full_name=current_user.full_name,
        phone=current_user.phone,
        date_of_birth=current_user.date_of_birth.isoformat() if current_user.date_of_birth else None,
        occupation=current_user.occupation,
        monthly_budget=current_user.monthly_budget,
        onboarding_completed=current_user.onboarding_completed or False,
        created_at=current_user.created_at if hasattr(current_user, 'created_at') else None,
        splitwise_user_id=current_user.splitwise_user_id,
        splitwise_last_sync_at=current_user.splitwise_last_sync_at
    )

@app.post("/api/profile/complete-onboarding")
def complete_onboarding(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Mark onboarding as completed"""
    current_user.onboarding_completed = True
    db.commit()
    
    return {"message": "Onboarding completed successfully"}


# ==========================================
# EXAMPLE CATEGORIES ROUTES
# ==========================================

@app.get("/api/categories/examples", response_model=List[ExampleCategoryResponse])
def get_example_categories(db: Session = Depends(get_db)):
    """
    Get list of example categories that users can choose from
    No authentication required - can be called during onboarding
    """
    examples = db.query(ExampleCategory).order_by(ExampleCategory.sort_order).all()
    return examples

@app.post("/api/categories/batch")
def create_categories_batch(
    batch: CategoryBatchCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Create multiple categories at once from example categories
    Used during onboarding when user selects multiple categories
    """
    if not batch.category_ids:
        raise HTTPException(status_code=400, detail="No categories selected")
    
    # Get example categories
    examples = db.query(ExampleCategory).filter(
        ExampleCategory.id.in_(batch.category_ids)
    ).all()
    
    if not examples:
        raise HTTPException(status_code=404, detail="No valid example categories found")
    
    created_count = 0
    skipped_count = 0
    
    for example in examples:
        # Check if user already has this category
        existing = db.query(Category).filter(
            Category.user_id == current_user.id,
            Category.name == example.name
        ).first()
        
        if existing:
            skipped_count += 1
            continue
        
        # Create category for user
        new_category = Category(
            user_id=current_user.id,
            name=example.name,
            color=example.color,
            icon=example.icon
        )
        db.add(new_category)
        created_count += 1
    
    db.commit()
    
    return {
        "message": "Categories created successfully",
        "created": created_count,
        "skipped": skipped_count,
        "total": len(batch.category_ids)
    }

# ==========================================
# CATEGORY MIGRATION ROUTE
# ==========================================

@app.post("/api/categories/migrate", response_model=CategoryMigrateResponse)
def migrate_category(
    migration: CategoryMigrateRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Migrate all transactions from one category to another
    Useful before deleting a category
    """
    # Check if both categories exist and belong to user
    from_category = db.query(Category).filter(
        Category.user_id == current_user.id,
        Category.name == migration.from_category_name
    ).first()
    
    to_category = db.query(Category).filter(
        Category.user_id == current_user.id,
        Category.name == migration.to_category_name
    ).first()
    
    if not from_category:
        raise HTTPException(status_code=404, detail=f"Source category '{migration.from_category_name}' not found")
    
    if not to_category:
        raise HTTPException(status_code=404, detail=f"Target category '{migration.to_category_name}' not found")
    
    # Use database function to migrate
    result = db.execute(
        "SELECT migrate_category_transactions(:user_id, :from_cat, :to_cat)",
        {
            "user_id": current_user.id,
            "from_cat": migration.from_category_name,
            "to_cat": migration.to_category_name
        }
    )
    
    affected_count = result.scalar()
    db.commit()
    
    return CategoryMigrateResponse(
        message=f"Successfully migrated {affected_count} transaction(s)",
        affected_count=affected_count,
        from_category=migration.from_category_name,
        to_category=migration.to_category_name
    )

# ==========================================
# UPDATED DELETE CATEGORY - With Migration Option
# ==========================================

@app.delete("/api/categories/{category_id}")
def delete_category_enhanced(
    category_id: int,
    migrate_to: Optional[str] = Query(None, description="Migrate transactions to this category before deleting"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Delete a category with optional transaction migration
    If migrate_to is provided, moves all transactions before deleting
    """
    category = db.query(Category).filter(
        Category.id == category_id,
        Category.user_id == current_user.id
    ).first()
    
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    
    # Check if category is used
    expense_count = db.query(Expense).filter(
        Expense.user_id == current_user.id,
        Expense.category == category.name
    ).count()
    
    # If category has expenses and no migration specified, reject
    if expense_count > 0 and not migrate_to:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete category '{category.name}' - it's used in {expense_count} expense(s). Please specify migrate_to parameter or reassign expenses first."
        )
    
    # If migration specified, migrate first
    if migrate_to:
        # Verify target category exists
        target = db.query(Category).filter(
            Category.user_id == current_user.id,
            Category.name == migrate_to
        ).first()
        
        if not target:
            raise HTTPException(status_code=404, detail=f"Migration target category '{migrate_to}' not found")
        
        # Migrate
        result = db.execute(
            "SELECT migrate_category_transactions(:user_id, :from_cat, :to_cat)",
            {
                "user_id": current_user.id,
                "from_cat": category.name,
                "to_cat": migrate_to
            }
        )
        migrated_count = result.scalar()
    
    # Delete category
    db.delete(category)
    db.commit()
    
    if migrate_to:
        return {
            "message": f"Category '{category.name}' deleted successfully after migrating {migrated_count} transaction(s) to '{migrate_to}'"
        }
    else:
        return {
            "message": f"Category '{category.name}' deleted successfully"
        }

# ==========================================
# iOS SHORTCUT INTEGRATION
# ==========================================

@app.get("/api/shortcut/generate")
def generate_ios_shortcut(
    current_user: User = Depends(get_current_user)
):
    """
    Generate iOS Shortcut configuration with user's auth token pre-filled
    
    Returns a .shortcut file content that can be imported to iOS
    """
    token = create_access_token(data={"sub": current_user.username})
    
    # iOS Shortcut configuration in plist format
    shortcut_config = {
        "api_url": f"{os.getenv('API_BASE_URL', 'https://webapp-expense.onrender.com')}/api/sms-parser/user/parse",
        "auth_token": token,
        "user_id": current_user.id,
        "username": current_user.username
    }
    
    return {
        "success": True,
        "shortcut_config": shortcut_config,
        "instructions": [
            "1. Copy the auth_token from above",
            "2. Open the iOS Shortcut link",
            "3. Paste the token when prompted",
            "4. Add shortcut to your library"
        ]
    }

@app.post("/api/sms-parser/user/parse")
async def parse_user_sms_authenticated(
    sms: str = Query(..., description="SMS text to parse"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Parse SMS and create pending transaction for authenticated user
    This endpoint is used by iOS Shortcuts
    
    Example: /api/sms-parser/user/parse?sms=Your%20A/c%20debited%20Rs.500
    Headers: Authorization: Bearer <token>
    
    Returns: Direct URL to approve transaction
    """
    from sms_parser_api import parse_sms_with_claude
    
    # Parse SMS
    parse_result = parse_sms_with_claude(sms)
    
    if not parse_result.get("success"):
        raise HTTPException(status_code=400, detail="Failed to parse SMS")
    
    data = parse_result["data"]
    
    # Create pending transaction
    pending = PendingTransaction(
        user_id=current_user.id,
        token=token_urlsafe(16),
        amount=data.get("amount"),
        description=data.get("merchant", "Unknown"),
        category=data.get("category", "Other"),
        date=data.get("date", datetime.now().strftime("%Y-%m-%d")),
        type="income" if data.get("transaction_type") == "credit" else "expense",
        status="pending"
    )
    
    db.add(pending)
    db.commit()
    db.refresh(pending)
    
    # Return URL for approval
    approval_url = f"{FRONTEND_URL}/add-expense/{pending.token}"
    
    return {
        "success": True,
        "url": approval_url,
        "parsed_data": data,
        "transaction_id": pending.id,
        "message": "Transaction parsed successfully. Open URL to approve."
    }

# ==========================================
# HEALTH CHECK
# ==========================================

@app.get("/")
def read_root():
    return {
        "app": "Expense Tracker API",
        "version": "3.0.0",
        "status": "running",
        "features": [
            "Authentication",
            "Password Reset",
            "Expenses CRUD",
            "Categories CRUD",
            "Export (CSV/Excel)",
            "Import (CSV/Excel)",
            "Pending Transactions",
            "SMS Parsing"
        ]
    }

@app.get("/health")
def health_check():
    return {"status": "healthy"}