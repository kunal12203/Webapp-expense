"""
Export/Import Module for Expense Tracker
Supports CSV and Excel export/import with optional email delivery
"""

from fastapi import APIRouter, Depends, HTTPException, File, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional, List
from datetime import datetime, date
from io import StringIO, BytesIO
import csv
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import os

# For Excel support
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  openpyxl not installed. Excel export will not be available.")
    print("   Install with: pip install openpyxl")

router = APIRouter()

# Email configuration (from main.py)
SMTP_HOST = os.getenv("SMTP_HOST")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD")
FROM_EMAIL = os.getenv("FROM_EMAIL", SMTP_USER)
FROM_NAME = os.getenv("FROM_NAME", "Expense Tracker")
EMAIL_ENABLED = bool(SMTP_HOST and SMTP_USER and SMTP_PASSWORD)

def export_to_csv(expenses: list) -> StringIO:
    """Export expenses to CSV format"""
    output = StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow(['Date', 'Amount', 'Category', 'Description', 'Type'])
    
    # Write data
    for expense in expenses:
        writer.writerow([
            expense.date.strftime("%Y-%m-%d") if hasattr(expense.date, 'strftime') else expense.date,
            expense.amount,
            expense.category,
            expense.description,
            expense.type
        ])
    
    output.seek(0)
    return output

def export_to_excel(expenses: list) -> BytesIO:
    """Export expenses to Excel format with formatting"""
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Excel export not available. Install openpyxl.")
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write header
    headers = ['Date', 'Amount', 'Category', 'Description', 'Type']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data
    for row, expense in enumerate(expenses, 2):
        ws.cell(row=row, column=1).value = expense.date.strftime("%Y-%m-%d") if hasattr(expense.date, 'strftime') else expense.date
        ws.cell(row=row, column=2).value = float(expense.amount)
        ws.cell(row=row, column=3).value = expense.category
        ws.cell(row=row, column=4).value = expense.description
        ws.cell(row=row, column=5).value = expense.type
        
        # Format amount as currency
        ws.cell(row=row, column=2).number_format = '$#,##0.00'
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12  # Date
    ws.column_dimensions['B'].width = 12  # Amount
    ws.column_dimensions['C'].width = 15  # Category
    ws.column_dimensions['D'].width = 30  # Description
    ws.column_dimensions['E'].width = 10  # Type
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def send_export_email(to_email: str, username: str, file_content: bytes, filename: str, file_type: str) -> bool:
    """Send export file via email"""
    if not EMAIL_ENABLED:
        print(f"⚠️  Email not configured. Cannot send export to {to_email}")
        return False
    
    try:
        # Create message
        msg = MIMEMultipart()
        msg['Subject'] = f'Your Expense Export - {filename}'
        msg['From'] = f"{FROM_NAME} <{FROM_EMAIL}>"
        msg['To'] = to_email
        
        # Email body
        body = f"""
        Hi {username},
        
        Your expense data export is attached to this email.
        
        File: {filename}
        Format: {file_type.upper()}
        Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
        
        You can open this file with Excel, Google Sheets, or any spreadsheet application.
        
        ---
        Expense Tracker Team
        """
        
        msg.attach(MIMEText(body, 'plain'))
        
        # Attach file
        attachment = MIMEBase('application', 'octet-stream')
        attachment.set_payload(file_content)
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition', f'attachment; filename="{filename}"')
        msg.attach(attachment)
        
        # Send email
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASSWORD)
            server.send_message(msg)
        
        print(f"✅ Export emailed to {to_email}")
        return True
        
    except Exception as e:
        print(f"❌ Failed to email export to {to_email}: {str(e)}")
        return False

def parse_csv_file(file_content: str) -> List[dict]:
    """Parse CSV file and return list of expense records"""
    expenses = []
    reader = csv.DictReader(StringIO(file_content))
    
    for row in reader:
        try:
            expense = {
                'date': row.get('Date', row.get('date', '')),
                'amount': float(row.get('Amount', row.get('amount', 0))),
                'category': row.get('Category', row.get('category', 'Uncategorized')),
                'description': row.get('Description', row.get('description', '')),
                'type': row.get('Type', row.get('type', 'expense')).lower()
            }
            
            # Validate required fields
            if not expense['date'] or not expense['amount']:
                continue
                
            # Validate type
            if expense['type'] not in ['expense', 'income']:
                expense['type'] = 'expense'
            
            expenses.append(expense)
        except (ValueError, KeyError) as e:
            print(f"⚠️  Skipping invalid row: {row} - {str(e)}")
            continue
    
    return expenses

def parse_excel_file(file_content: bytes) -> List[dict]:
    """Parse Excel file and return list of expense records"""
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Excel import not available. Install openpyxl.")
    
    expenses = []
    wb = openpyxl.load_workbook(BytesIO(file_content))
    ws = wb.active
    
    # Get header row
    headers = [cell.value for cell in ws[1]]
    
    # Process data rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            row_dict = dict(zip(headers, row))
            
            expense = {
                'date': str(row_dict.get('Date', row_dict.get('date', ''))),
                'amount': float(row_dict.get('Amount', row_dict.get('amount', 0))),
                'category': str(row_dict.get('Category', row_dict.get('category', 'Uncategorized'))),
                'description': str(row_dict.get('Description', row_dict.get('description', ''))),
                'type': str(row_dict.get('Type', row_dict.get('type', 'expense'))).lower()
            }
            
            # Validate required fields
            if not expense['date'] or not expense['amount']:
                continue
                
            # Validate type
            if expense['type'] not in ['expense', 'income']:
                expense['type'] = 'expense'
            
            expenses.append(expense)
        except (ValueError, KeyError, TypeError) as e:
            print(f"⚠️  Skipping invalid row: {row} - {str(e)}")
            continue
    
    return expenses

# Export endpoints will be added to main.py
# Import endpoint will be added to main.py